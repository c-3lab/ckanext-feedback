from io import BytesIO
from unittest.mock import patch

import pytest
from ckan.lib import api_token as api_token_lib
from flask import Response
from openpyxl import load_workbook

from ckanext.feedback.controllers.api.moral_check_log import (
    create_moral_check_log_excel_response,
    create_moral_check_log_excel_workbook,
    download_moral_check_log,
)


@pytest.mark.db_test
class TestMoralCheckLog:
    @patch(
        'ckanext.feedback.controllers.api.moral_check_log.'
        'resource_comment_service.get_resource_comment_moral_check_logs'
    )
    @patch(
        'ckanext.feedback.controllers.api.moral_check_log.'
        'utilization_detail_service.get_utilization_comment_moral_check_logs'
    )
    def test_create_moral_check_log_excel_workbook(
        self,
        mock_get_utilization_comment_moral_check_logs,
        mock_get_resource_comment_moral_check_logs,
        resource_comment_moral_check_log,
        utilization_comment_moral_check_log,
    ):
        mock_get_resource_comment_moral_check_logs.return_value = [
            resource_comment_moral_check_log
        ]
        mock_get_utilization_comment_moral_check_logs.return_value = [
            utilization_comment_moral_check_log
        ]

        output = create_moral_check_log_excel_workbook(False)

        workbook = load_workbook(filename=BytesIO(output.getvalue()))

        assert "MoralCheckLog" in workbook.sheetnames

        sheet = workbook['MoralCheckLog']

        headers = [cell.value for cell in sheet[1]]
        assert headers == [
            'id',
            'type',
            'resource_or_utilization_id',
            'action',
            'input_comment',
            'suggested_comment',
            'output_comment',
            'timestamp',
        ]

        data = list(sheet.iter_rows(values_only=True))
        assert data[1] == (
            resource_comment_moral_check_log.id,
            'resource',
            resource_comment_moral_check_log.resource_id,
            resource_comment_moral_check_log.action.name,
            resource_comment_moral_check_log.input_comment,
            resource_comment_moral_check_log.suggested_comment,
            resource_comment_moral_check_log.output_comment,
            resource_comment_moral_check_log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
        )
        assert data[2] == (
            utilization_comment_moral_check_log.id,
            'utilization',
            utilization_comment_moral_check_log.utilization_id,
            utilization_comment_moral_check_log.action.name,
            utilization_comment_moral_check_log.input_comment,
            utilization_comment_moral_check_log.suggested_comment,
            utilization_comment_moral_check_log.output_comment,
            utilization_comment_moral_check_log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
        )

    @patch(
        'ckanext.feedback.controllers.api.moral_check_log.'
        'resource_comment_service.get_resource_comment_moral_check_logs'
    )
    @patch(
        'ckanext.feedback.controllers.api.moral_check_log.'
        'utilization_detail_service.get_utilization_comment_moral_check_logs'
    )
    def test_create_moral_check_log_excel_workbook_with_separation(
        self,
        mock_get_utilization_comment_moral_check_logs,
        mock_get_resource_comment_moral_check_logs,
        resource_comment_moral_check_log,
        utilization_comment_moral_check_log,
    ):
        mock_get_resource_comment_moral_check_logs.return_value = [
            resource_comment_moral_check_log
        ]
        mock_get_utilization_comment_moral_check_logs.return_value = [
            utilization_comment_moral_check_log
        ]

        output = create_moral_check_log_excel_workbook(True)

        workbook = load_workbook(filename=BytesIO(output.getvalue()))

        assert "ResourceCommentMoralCheckLog" in workbook.sheetnames
        assert "UtilizationCommentMoralCheckLog" in workbook.sheetnames

        resource_sheet = workbook['ResourceCommentMoralCheckLog']
        utilization_sheet = workbook['UtilizationCommentMoralCheckLog']

        resource_headers = [cell.value for cell in resource_sheet[1]]
        assert resource_headers == [
            'id',
            'resource_id',
            'action',
            'input_comment',
            'suggested_comment',
            'output_comment',
            'timestamp',
        ]
        utilization_headers = [cell.value for cell in utilization_sheet[1]]
        assert utilization_headers == [
            'id',
            'utilization_id',
            'action',
            'input_comment',
            'suggested_comment',
            'output_comment',
            'timestamp',
        ]

        resource_data = list(resource_sheet.iter_rows(values_only=True))
        utilization_data = list(utilization_sheet.iter_rows(values_only=True))

        assert resource_data[1] == (
            resource_comment_moral_check_log.id,
            resource_comment_moral_check_log.resource_id,
            resource_comment_moral_check_log.action.name,
            resource_comment_moral_check_log.input_comment,
            resource_comment_moral_check_log.suggested_comment,
            resource_comment_moral_check_log.output_comment,
            resource_comment_moral_check_log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
        )
        assert utilization_data[1] == (
            utilization_comment_moral_check_log.id,
            utilization_comment_moral_check_log.utilization_id,
            utilization_comment_moral_check_log.action.name,
            utilization_comment_moral_check_log.input_comment,
            utilization_comment_moral_check_log.suggested_comment,
            utilization_comment_moral_check_log.output_comment,
            utilization_comment_moral_check_log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
        )

    @patch(
        'ckanext.feedback.controllers.api.moral_check_log.'
        'create_moral_check_log_excel_workbook'
    )
    def test_create_moral_check_log_excel_response(
        self, mock_create_moral_check_log_excel_workbook
    ):
        mock_output = BytesIO(b'test data')

        mock_create_moral_check_log_excel_workbook.return_value = mock_output

        response = create_moral_check_log_excel_response(False)

        assert isinstance(response, Response)
        assert response.status_code == 200
        assert (
            response.mimetype
            == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        assert (
            response.headers['Content-Disposition']
            == 'attachment; filename="moral_check_log.xlsx"'
        )
        assert response.data == b'test data'

    @patch(
        'ckanext.feedback.controllers.api.moral_check_log.'
        'create_moral_check_log_excel_workbook'
    )
    def test_create_moral_check_log_excel_response_with_separation(
        self, mock_create_moral_check_log_excel_workbook
    ):
        mock_output = BytesIO(b'test data')

        mock_create_moral_check_log_excel_workbook.return_value = mock_output

        response = create_moral_check_log_excel_response(True)

        assert isinstance(response, Response)
        assert response.status_code == 200
        assert (
            response.mimetype
            == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        assert (
            response.headers['Content-Disposition']
            == 'attachment; filename="moral_check_log_separation.xlsx"'
        )
        assert response.data == b'test data'

    @pytest.mark.usefixtures('with_request_context')
    @patch('ckanext.feedback.controllers.api.moral_check_log.request.headers.get')
    @patch(
        'ckanext.feedback.controllers.api.moral_check_log.'
        'AuthTokenHandler.validate_api_token'
    )
    @patch(
        'ckanext.feedback.controllers.api.moral_check_log.'
        'AuthTokenHandler.decode_api_token'
    )
    @patch(
        'ckanext.feedback.controllers.api.moral_check_log.'
        'user_service.get_user_by_token_id'
    )
    @patch(
        'ckanext.feedback.controllers.api.moral_check_log.'
        'AuthTokenHandler.check_sysadmin'
    )
    @patch('ckanext.feedback.controllers.api.moral_check_log.request.args.get')
    @patch(
        'ckanext.feedback.controllers.api.moral_check_log.'
        'create_moral_check_log_excel_response'
    )
    def test_download_moral_check_log(
        self,
        mock_create_moral_check_log_excel_response,
        mock_request_args_get,
        mock_check_sysadmin,
        mock_get_user_by_token_id,
        mock_decode_api_token,
        mock_validate_api_token,
        mock_request_headers_get,
        user,
        api_token,
    ):
        mock_request_headers_get.return_value = api_token['token']
        mock_validate_api_token.return_value = None
        mock_decode_api_token.return_value = api_token_lib.decode(api_token['token'])[
            'jti'
        ]
        mock_get_user_by_token_id.return_value = user
        mock_check_sysadmin.return_value = None
        mock_request_args_get.return_value = False
        mock_create_moral_check_log_excel_response.return_value = Response(
            b'test data',
            mimetype=(
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            ),
            headers={
                'Content-Disposition': 'attachment; filename="moral_check_log.xlsx"'
            },
        )

        response = download_moral_check_log()

        assert response.status_code == 200
        assert (
            response.mimetype
            == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        assert (
            response.headers['Content-Disposition']
            == 'attachment; filename="moral_check_log.xlsx"'
        )
        assert response.data == b'test data'
